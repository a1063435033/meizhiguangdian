import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# 读取Excel文件并进行数据处理
def process_data(file_path):
    df = pd.read_excel(file_path)

    # 创建一个ExcelWriter对象，用于保存多个工作表
    with pd.ExcelWriter('output.xlsx') as writer:
        # 遍历每一组两列数据
        for i in range(0, len(df.columns), 2):
            # 获取当前组的两列数据
            time_col = df.columns[i]
            ppfd_col = df.columns[i + 1]
            
            # 提取这两列数据
            subset = df[[time_col, ppfd_col]].copy()  # 使用 copy() 避免 SettingWithCopyWarning
            
            # 转换时间格式
            def convert_time(time_str):
                if pd.isna(time_str):  # 检查是否为空值
                    return ""  # 如果是空值，返回空字符串
                try:
                    # 解析时间并格式化为 HH:MM:SS
                    return datetime.strptime(str(time_str), "%Y-%m-%d:%H:%M:%S").strftime("%H:%M:%S")
                except ValueError:
                    return ""  # 如果时间格式无效，返回空字符串
            
            subset[time_col] = subset[time_col].apply(convert_time)
            
            # 获取列名作为工作表名称
            sheet_name = time_col  # 使用时间列的列名作为工作表名称
            
            # 保存到新的工作表中
            subset.to_excel(writer, sheet_name=sheet_name, index=False)

    print("数据处理完成，已保存到output.xlsx文件中。")
    return 'output.xlsx'  # 返回处理后的文件路径

# 在Excel中嵌入图表
def embed_charts_in_excel(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 动态获取有效数据范围（跳过空行）
        data = list(ws.values)
        headers = data[0]
        df = pd.DataFrame(data[1:], columns=headers).dropna(how='all')  # 过滤空行
        
        # 计算数据范围
        last_data_row = len(df) + 1  # 数据结束行（表头占1行）
        
        # 时间列（X轴）范围
        time_ref = Reference(ws, 
                           min_col=1, 
                           min_row=2,  # 从第2行开始（跳过表头）
                           max_col=1, 
                           max_row=last_data_row + 1)  # 包含最后一行数据
        
        # PPFD列（Y轴）范围
        ppfd_ref = Reference(ws, 
                            min_col=2, 
                            min_row=1,  # 从第2行开始（跳过表头）
                            max_col=2, 
                            max_row=last_data_row)  # 包含最后一行数据
        
        # 创建折线图
        chart = LineChart()
        chart.title = f"PPFD趋势 - {sheet_name}"
        chart.style = 13
        chart.y_axis.title = "PPFD (μmol/m²/s)"
        chart.x_axis.title = "时间"
        
        # 添加数据系列
        chart.add_data(ppfd_ref, titles_from_data=True)
        chart.set_categories(time_ref)
        
        # 设置图表位置
        ws.add_chart(chart, "D2")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18

    # 保存文件
    new_file = file_path.replace(".xlsx", "_with_charts.xlsx")
    wb.save(new_file)
    print(f"文件已保存：{new_file}")

# 主函数
def main():
    # 原始文件路径
    input_file = r"D:\testwu\demo\combined_data(1).xlsx"
    
    # 处理数据并生成新的Excel文件
    processed_file = process_data(input_file)
    
    # 在生成的Excel文件中嵌入图表
    embed_charts_in_excel(processed_file)

# 执行主函数
if __name__ == "__main__":
    main()