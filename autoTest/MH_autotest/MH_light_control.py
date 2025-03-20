
import serial
import time
import json
import yaml
import os
import re
import utils.light_commands
from datetime import datetime, timedelta
from openpyxl import Workbook
from page.fileOperations import FileOperations
from serial_data_parser import SerialDataParser

class LightControl(FileOperations):
    def __init__(self, filepath):
        super().__init__(filepath)
        self.timeout_seconds = 7
        config_path = r"D:\meizhiguangdian\autoTest\MH_autotest\config\config.yaml"
        with open(config_path, 'r', encoding='utf-8') as file:
            self.config_yaml = yaml.safe_load(file)
        self.st = self.config_yaml['autoParams']['st']
        self.et = self.config_yaml['autoParams']['et']
        self.darkenT = self.config_yaml['autoParams']['darkenT']
        self.offT = self.config_yaml['autoParams']['offT']
        self.pid = self.config_yaml['device']['pid']
        self.ser_report = self.config_yaml['device']['ser']
        self.dim_min = self.config_yaml['ppfdParams']['dim_min']
        self.dim_max = self.config_yaml['ppfdParams']['dim_max']
        self.sunTime = self.config_yaml['ppfdParams']['sunTime']
    
    def manual_light_adjust_mode(self, ser, dim):
        commands = utils.light_commands.get_manual_light_adjust_mode_Cmd(self.pid, 1, dim)
        ser.write(commands.encode('utf-8'))
        time.sleep(0.5)
        buffer = SerialDataParser.read_serial_succeed_data(ser, self.timeout_seconds)
        if buffer["code"] == 200:
            self.write_log(f"发送手动调节灯光模式命令成功: {commands}", 'manual_light_adjust_mode.txt')
        else:
            self.write_log(f"6S内没有接受到指令响应，发送失败", 'manual_light_adjust_mode.txt')
            return None
        start_time = time.time()
        while time.time() - start_time < self.timeout_seconds:  # 在6秒内持续读取
            if ser.in_waiting > 0:
                raw_data = ser.readline().decode().rstrip()
                # print(raw_data)
                # 读取一行数据    
                first_json_end = raw_data.find('}{') + 1  # 找到'}{'的位置，并调整以包含第一个'}'
                # # 分割字符串
                # first_json_str = buffer[:first_json_end]
                second_json_str = raw_data[first_json_end:]
                # 解析JSON字符串为字典
                # first_dict = json.loads(first_json_str)
                second_dict = json.loads(second_json_str)
                light_data = second_dict.get('params', {}).get('light')
                if light_data['dim'] == dim:
                    self.write_log(f"调节亮度成功, 亮度dim值为：{light_data['dim']}", 'manual_light_adjust_mode.txt')
                    return second_dict['params']['sensor']['ppfd']
                    
                else:
                    self.write_log(f"调节亮度失败, 亮度dim值为：{light_data['dim']}", 'manual_light_adjust_mode.txt')
        return None
      
    def manual_light_adjust_mode_test(self, dimensions, ser):
        for dim in dimensions:
            manual_dim = self.manual_light_adjust_mode(ser,dim)
            if manual_dim is not None:
                ppfd_num = self.config_yaml['dimconfig'][f'dim{dim}']         
                difference = manual_dim - ppfd_num
                status = "通过" if -self.config_yaml['margins']['margin'] <= difference <= self.config_yaml['margins']['margin'] else "失败"
                log_msg = f"手动测试{status}: 设置{dim}%亮度，ppfd值为{manual_dim}, 差值为{difference}"
                self.write_log(log_msg, 'manual_light_adjust_mode_report.txt')
            else:
                self.write_log(f'manual_dim得值为：{manual_dim},请检查代码', 'test_report.txt')

    def auto_light_adjust_mode(self, ser, pid, st, et, dim, darkenT, offT, sunTime):
        ppfd_list = []
        current_time_list = []
        commands = utils.light_commands.get_auto_light_adjust_mode_Cmd(pid, st, et, dim, darkenT, offT, sunTime)
        ser.write(commands.encode('utf-8'))
        time.sleep(0.5)
        buffer = SerialDataParser.read_serial_succeed_data(ser, self.timeout_seconds)
        if buffer["code"] == 200:
            self.write_log(f"发送手动调节灯光模式命令成功: {commands}", 'manual_light_adjust_mode.txt')
        else:
            self.write_log(f"6S内没有接受到指令响应，发送失败", 'manual_light_adjust_mode.txt')
            return None
        
        while True:
            if ser.in_waiting > 0:
                raw_data = SerialDataParser.split_json_pairs(ser)
                first_dict = raw_data[0]
                second_dict = raw_data[1]
                minutes_since_midnights = first_dict['params']['device_info']['timeinfo']
                dt = datetime.strptime(minutes_since_midnights.rsplit('-', 1)[0], '%Y-%m-%d:%H:%M:%S')
                # 计算从当天开始到现在经过了多少分钟
                midnight = dt.replace(hour=0, minute=0, second=0, microsecond=0)
                minutes_passed = (dt - midnight).total_seconds() / 60

                if minutes_passed >= st:
                    self.write_log(second_dict['params']['sensor']['ppfd'], 'auto_light_adjust_mode_log.txt')
                    self.write_log(minutes_since_midnights.rsplit('-', 1)[0], 'auto_light_adjust_mode_log.txt')
                    ppfd_list.append(second_dict['params']['sensor']['ppfd'])

                    current_time_list.append(minutes_since_midnights.rsplit('-', 1)[0])
                    if second_dict['params']['sensor']['ppfd'] == 0 and minutes_passed >= et:
                        self.write_log(second_dict['params']['sensor']['ppfd'], 'auto_light_adjust_mode_log.txt')
                        self.write_log(minutes_since_midnights.rsplit('-', 1)[0], 'auto_light_adjust_mode_log.txt')
                        return ppfd_list, current_time_list
                    elif second_dict['params']['sensor']['ppfd'] == -1:
                        ppfd = second_dict['params']['sensor']['ppfd']
                        self.write_log(f'获取不到PPFD值，PPFD值为：{ppfd}，请检查PPFD传感器问题', 'auto_light_adjust_mode_log.txt')
                        return None
                    
                if minutes_passed >= et and second_dict['params']['sensor']['ppfd'] != 0:
                    ppfd = second_dict['params']['sensor']['ppfd']
                    self.write_log(f'当前时间以超过{et}，PPFD：{ppfd}，灯光未关闭', 'auto_light_adjust_mode_log.txt')
                    return None
                
    def auto_light_adjust_mode_test(self, ser):
        testReport = {}
        nowTime = SerialDataParser.read_serial_data_get_nowTime(ser)
        st = int(nowTime[0]) + 2
        et = st + 3
        for sunTime in range(1, 5): # 循环遍历一个月中的每一天
            # 跨天
            if et >=1430:
                st = 0
                et = st + 2 * sunTime + 1
            hours, minutes = st // 60, st % 60
            end_hours, end_minutes = et // 60, et % 60
            new_time = f"{hours:02}:{minutes:02}"
            end_time = f"{end_hours:02}:{end_minutes:02}"
            self.write_log(f'开始测试,开始时间为：{new_time},结束时间为：{end_time},日出日落时间为：{sunTime}', 'auto_light_mode.txt')
            data = self.auto_light_adjust_mode(ser, self.pid, st, et, 10, self.darkenT, self.offT, sunTime)
            if data is None:
                self.write_log(f"Data is None at sunTime={sunTime}. Skipping this iteration.")
                continue  # 如果 data 是 None，则跳过当前迭代
            # 更新下一天的开始时间为今天的结束时间加1
            st = et + 1
            # 根据当前的sunTime计算当天的结束时间
            et = st + 2 * (sunTime + 1) + 1
            testReport[f'日出日落第{sunTime}分钟时间'] = data[1]
            testReport[f'日出日落第{sunTime}分钟ppfd值'] = data[0]

            print('-----------------------------------------------------------------')
            time.sleep(5)
        # 创建一个新的工作簿和工作表
        wb = Workbook()
        ws = wb.active

        # 添加标题行
        column_names = list(testReport.keys())
        for col_num, column_name in enumerate(column_names, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        # 添加数据
        max_length = max(len(v) for v in testReport.values())  # 获取最长列的长度
        for row_num in range(2, max_length + 2):  # 从第二行开始添加数据
            for col_num, column_name in enumerate(column_names, 1):
                if row_num - 2 < len(testReport[column_name]):
                    ws.cell(row=row_num, column=col_num, value=testReport[column_name][row_num - 2])

        # 保存Excel文件
        wb.save('combined_data.xlsx')
    def ppfd_light_adjust_mode(self, ser, ppfd):
        nowTime = SerialDataParser.read_serial_data_get_nowTime(ser)
        st = int(nowTime[0]) + 2
        et = st + 10
        commands = utils.light_commands.get_ppfd_light_adjust_mode_Cmd(self.pid, st, et, self.dim_min, self.dim_max, ppfd, self.darkenT, self.offT, self.sunTime)
        ser.write(commands.encode('utf-8'))
        time.sleep(0.5)
        buffer = SerialDataParser.read_serial_succeed_data(ser, self.timeout_seconds)
        if buffer["code"] == 200:
            self.write_log(f"发送手动调节灯光模式命令成功: {commands}", 'manual_light_adjust_mode.txt')
        else:
            self.write_log(f"6S内没有接受到指令响应，发送失败", 'manual_light_adjust_mode.txt')
            return None
        while True:
            if ser.in_waiting > 0:
                # 读取一行数据
                raw_data = ser.readall().decode().rstrip() 
                first_json_end = raw_data.find('}{') + 1  # 找到'}{'的位置，并调整以包含第一个'}'
                # # 分割字符串
                first_json_str = raw_data[:first_json_end]
                second_json_str = raw_data[first_json_end:]
                # 解析JSON字符串为字典
                first_dict = json.loads(first_json_str)
                second_dict = json.loads(second_json_str)
                minutes_since_midnight = first_dict['params']['device_info']['timeinfo']
                dt = datetime.strptime(minutes_since_midnight.rsplit('-', 1)[0], '%Y-%m-%d:%H:%M:%S')
                # 计算从当天开始到现在经过了多少分钟
                midnight = dt.replace(hour=0, minute=0, second=0, microsecond=0)
                minutes_passed = (dt - midnight).total_seconds() / 60
                dim = second_dict['params']['light']['dim']
                ppfd = second_dict['params']['sensor']['ppfd']
                if minutes_passed >= st and  minutes_passed <= st + 1:
                    if second_dict['params']['light']['dim'] == self.dim_min:
                        self.write_log(f'设备dim值为:{dim},自定义dim_min值为{self.dim_min}', 'ppfd_light_adjust_mode.txt')

                    else:
                        self.write_log(f'一分钟内，设备dim值不等于dim_min值，{dim}、{self.dim_min}', 'ppfd_light_adjust_mode.txt')
                        return False
                elif minutes_passed > st + 1:
                    self.write_log(second_dict['params']['sensor']['ppfd'], 'ppfd_light_adjust_mode.txt')
                    self.write_log(second_dict['params']['light']['dim'], 'ppfd_light_adjust_mode.txt')
                    if second_dict['params']['sensor']['ppfd'] > self.dim_max + 3:
                        self.write_log(f'当前ppfd超过最大值，ppfd值为：{ppfd}，最大值为：{self.dim_max}', 'ppfd_light_adjust_mode.txt')
                        return False
                    # elif second_dict['params']['sensor']['ppfd']
                elif minutes_passed >= et:
                    print('测试结束')
                    return True
                
    def run_light_control_test(self):
        print('开始手动调节灯光亮度测试')
        print('-----------------------------------------------------------------')
        ser = serial.Serial(self.ser_report, 115200, timeout=1)
        ppfd = self.config_yaml['ppfdParams']['ppfd']
        # 正向测试: 从start_dim到100，每步增加10%
        ascending_sequence = range(10, 110, 10)
        self.manual_light_adjust_mode_test(ascending_sequence, ser)
        # 反向测试: 从100到start_dim，每步减少10%
        descending_sequence = range(100, 10-10, -10)
        self.manual_light_adjust_mode_test(descending_sequence, ser)
        print('手动调节灯光亮度测试结束')
        print('-----------------------------------------------------------------')
        print('开始自动灯光调节测试')
        print('-----------------------------------------------------------------')
        self.auto_light_adjust_mode_test(ser)
        print('自动调节灯光亮度测试结束')
        print('-----------------------------------------------------------------')
        print('开始PPFD模式灯光调节测试')
        print('-----------------------------------------------------------------')
        # for i in range(1, 6):
        #     self.ppfd_light_adjust_mode(ser, ppfd)
        #     # ppfd -= 500
        #     time.sleep(5)
        print('控制器灯光控制测试结束')
        print('-----------------------------------------------------------------')







                


                


                        
                
        



                        

                    
# # 调用方法
# if __name__ == "__main__":
#     # 获取当前路径并设置日志文件夹路径
#     current_path = os.getcwd()
#     now_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # 当前时间
#     result_path = os.path.join(current_path, f"MH控制器测试日志{now_time}")
#     if not os.path.exists(result_path):
#         os.makedirs(result_path)
#     log_path = result_path
#     a = LightControl(log_path)
#     a. run_light_control_test()
