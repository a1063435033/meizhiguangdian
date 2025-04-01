import time
import json
import os
import re
import yaml
import serial
import utils.light_commands
from page.fileOperations import FileOperations
from utils.serial_data_parser import SerialDataParser
from datetime import datetime, timedelta
from openpyxl import Workbook


class PowerStripControl(FileOperations):
    def __init__(self, filepath):
        super().__init__(filepath)
        self.timeout_seconds = 7
        config_path = r"D:\meizhiguangdian_test\meizhiguangdian\autoTest\MH_autotest\config\powerStrip_config.yaml"
        with open(config_path, 'r', encoding='utf-8') as file:
            self.config_yaml = yaml.safe_load(file)
        self.times = self.config_yaml['switch_times']
        self.pid = self.config_yaml['device']['pid']

    def check_iHub4Set_switch_status(self, ser, start_time, ihub, on):
        while time.time() - start_time < self.timeout_seconds:
            if ser.in_waiting > 0:
                raw_data = ser.readline().decode().rstrip()
                first_json_end = raw_data.find('}{') + 1  # 找到'}{'的位置，并调整以包含第一个'}'
                # # 分割字符串
                # first_json_str = buffer[:first_json_end]
                second_json_str = raw_data[first_json_end:]
                # 解析JSON字符串为字典
                # first_dict = json.loads(first_json_str)
                second_dict = json.loads(second_json_str)
                light_data = second_dict.get('params', {}).get("iHub")[ihub]
                if light_data['on'] == on:
                    return True
                else:
                    self.write_log(f"当前发送状态为：{on}，软件上报状态为{light_data['on']}，请检查手动模式开关孔位操作", 'iHub4Set_switch_manual_mode.txt')
                    return False

    def iHub4Set_switch_manual_mode(self, ser, pid):
        ihubs = ["uv", "ir", "rb"]
        def control_socket(ihub):
            for i in range(self.times):
                sendSwitchManualCmd = utils.light_commands.get_iHub4Set_switch_manual_mode_Cmd(ser, ihub, pid, 1)
                if not sendSwitchManualCmd:
                    self.write_log("发送指令失败，请检查串口", 'iHub4Set_switch_manual_mode.txt')
                    return False  # 或者采取其他适当的操作
                start_time = time.time()
                if self.check_iHub4Set_switch_status(ser, start_time, ihub, 1):
                    print(f'第{i+1}次打开{ihub}孔位')               
                sendSwitchManualCmd = utils.light_commands.get_iHub4Set_switch_manual_mode_Cmd(ser, ihub, pid, 0)
                if not sendSwitchManualCmd:
                    self.write_log("发送指令失败，请检查串口", 'iHub4Set_switch_manual_mode.txt')
                    return False  # 或者采取其他适当的操作
                start_time = time.time()
                if self.check_iHub4Set_switch_status(ser, start_time, ihub, 0):
                    print(f'第{i+1}次关闭{ihub}孔位')
            return True
        for ihub in ihubs:
            if control_socket(ihub):
                self.write_log(f"四孔排插{ihub}位开关测试通过", 'iHub4Set_switch_manual_mode.txt')
            else:
                self.write_log(f"四孔排插{ihub}位开关测试失败", 'iHub4Set_switch_manual_mode.txt')

    def iHub4Set_switch_autoMod_mode(self, ser, pid):
        rd = 60
        ed = 60
        times = 99
        ihubs = ["uv", "ir", "rb"]
        nowTime = SerialDataParser.read_serial_data_get_nowTime(ser)
        date_time_part, tz_offset = nowTime[1].rsplit('-', 1)
        tz_offset = '-' + tz_offset  # 添加负号，确保时区偏移正确
        dt = datetime.strptime(date_time_part, '%Y-%m-%d:%H:%M:%S')
        new_dt = dt + timedelta(minutes=2)
        st = int(nowTime[0]) + 2
        for ihub in ihubs:
            sendSwitchManualCmd = utils.light_commands.get_iHub4Set_switch_autoMod_mode_Cmd(ser, pid, ihub, st, rd, ed, times)
            if not sendSwitchManualCmd:
                self.write_log("发送指令失败，请检查串口", 'iHub4Set_switch_autoMod_mode.txt')
                return False
            self.write_log(f"四孔排插{ihub}位循环测试开始，开始时间为：{new_dt}", 'iHub4Set_switch_manual_mode.txt')

    def iHub4Set_light_manual_mode(self, ser, dim):
        sendLightManualModeCmd = utils.light_commands.get_iHub4Set_light_manual_mode_Cmd(ser, self.pid, 1, dim)
        if not sendLightManualModeCmd:
            self.write_log("发送指令失败，请检查串口", 'iHub4Set_light_manual_mode.txt')
            return False
        light_data, second_dict= SerialDataParser.get_serial_data(ser)
        deviceDim = second_dict['params']['iHub']['light']['dim']
        if deviceDim == dim:
            self.write_log(f"调节亮度成功, 亮度dim值为：{deviceDim}", 'iHub4Set_light_manual_mode.txt')
            return second_dict['params']['sensor']['ppfd']
                    
        else:
            self.write_log(f"调节亮度失败, 亮度dim值为：{deviceDim}", 'iHub4Set_light_manual_mode.txt')
            return False
        
    def iHub4Set_light_manual_mode_test(self, ser):
        ascending_sequence = range(10, 110, 10)
        descending_sequence = range(100, 10-10, -10)
        for dim in ascending_sequence:
            manual_dim = self.iHub4Set_light_manual_mode(ser, dim)
            if manual_dim:
                ppfd_num = self.config_yaml['dimconfig'][f'dim{dim}']
                difference = manual_dim - ppfd_num
                status = "通过" if -self.config_yaml['margins']['margin'] <= difference <= self.config_yaml['margins']['margin'] else "失败"
                log_msg = f"手动测试{status}: 设置{dim}%亮度，ppfd值为{manual_dim}, 差值为{difference}, 设定值为{ppfd_num}"
                self.write_log(log_msg, 'iHub4Set_light_manual_mode.txt')
            else:
                self.write_log(f'manual_dim得值为：{manual_dim},请检查代码', 'iHub4Set_light_manual_mode.txt')
        for dim in descending_sequence:
            manual_dim = self.iHub4Set_light_manual_mode(ser, dim)
            if manual_dim:
                ppfd_num = self.config_yaml['dimconfig'][f'dim{dim}']
                difference = manual_dim - ppfd_num
                status = "通过" if -self.config_yaml['margins']['margin'] <= difference <= self.config_yaml['margins']['margin'] else "失败"
                log_msg = f"手动测试{status}: 设置{dim}%亮度，ppfd值为{manual_dim}, 差值为{difference}, 设定值为{ppfd_num}"
                self.write_log(log_msg, 'iHub4Set_light_manual_mode.txt')
            else:
                self.write_log(f'manual_dim得值为：{manual_dim},请检查代码', 'iHub4Set_light_manual_mode.txt')

    def iHub4Set_light_auto_mode(self, ser, pid, st, et, dim, darkenT, offT, sunTime):
        ppfd_list = []
        current_time_list = []
        sendAutoModCmd = utils.light_commands.get_iHub4Set_light_auto_mode_Cmd(ser, pid, st, et, dim, darkenT, offT, sunTime)
        if not sendAutoModCmd:
            self.write_log("发送指令失败，请检查串口", 'iHub4Set_light_auto_mode.txt')
            return False
        while True:
            if ser.in_waiting > 0:
                raw_data = SerialDataParser.split_json_pairs(ser)
                first_dict = raw_data[0]
                second_dict = raw_data[1]
                minutes_since_midnights = first_dict['params']['device_info']['timeinfo']
                dt = datetime.strptime(minutes_since_midnights.rsplit('-', 1)[0], '%Y-%m-%d:%H:%M:%S')
                # 计算从当天开始到现在经过了多少分钟
                midnight = dt.replace(hour=0, minute=0, second=0, microsecond=0)
                minutes_passed = (dt - midnight).total_seconds() / 60

                if minutes_passed >= st:
                    self.write_log(second_dict['params']['sensor']['ppfd'], 'iHub4Set_light_auto_mode.txt')
                    self.write_log(minutes_since_midnights.rsplit('-', 1)[0], 'iHub4Set_light_auto_mode.txt')
                    ppfd_list.append(second_dict['params']['sensor']['ppfd'])

                    current_time_list.append(minutes_since_midnights.rsplit('-', 1)[0])
                    if second_dict['params']['sensor']['ppfd'] == 0 and minutes_passed >= et:
                        return ppfd_list, current_time_list
                    elif second_dict['params']['sensor']['ppfd'] == -1:
                        ppfd = second_dict['params']['sensor']['ppfd']
                        self.write_log(f'获取不到PPFD值，PPFD值为：{ppfd}，请检查PPFD传感器问题', 'iHub4Set_light_auto_mode.txt')
                        return None
                    
                if minutes_passed >= et + 1 and second_dict['params']['sensor']['ppfd'] != 0:
                    ppfd = second_dict['params']['sensor']['ppfd']
                    self.write_log(f'当前时间以超过{et}，PPFD：{ppfd}，灯光未关闭', 'iHub4Set_light_auto_mode.txt')
                    return None
                
    def iHub4Set_light_auto_mode_test(self, ser):
            testReport = {}
            nowTime = SerialDataParser.read_serial_data_get_nowTime(ser)
            st = int(nowTime[0]) + 2
            et = st + 3
            for sunTime in range(1, 3): # 循环遍历一个月中的每一天
                # 跨天
                if et >=1430:
                    st = 0
                    et = st + 2 * sunTime + 1
                hours, minutes = st // 60, st % 60
                end_hours, end_minutes = et // 60, et % 60
                new_time = f"{hours:02}:{minutes:02}"
                end_time = f"{end_hours:02}:{end_minutes:02}"
                self.write_log(f'开始测试,开始时间为：{new_time},结束时间为：{end_time},日出日落时间为：{sunTime}', 'iHub4Set_light_auto_mode.txt')
                data = self.iHub4Set_light_auto_mode(ser, self.pid, st, et, 50, 45, 50, sunTime)
                if data is None:
                    self.write_log(f"Data is None at sunTime={sunTime}. Skipping this iteration.")
                    continue  # 如果 data 是 None，则跳过当前迭代
                # 更新下一天的开始时间为今天的结束时间加1
                st = et + 1
                # 根据当前的sunTime计算当天的结束时间
                et = st + 2 * (sunTime + 1) + 1
                testReport[f'日出日落第{sunTime}分钟时间'] = data[1]
                testReport[f'日出日落第{sunTime}分钟ppfd值'] = data[0]

                print('-----------------------------------------------------------------')
                time.sleep(5)
            # 创建一个新的工作簿和工作表
            wb = Workbook()
            ws = wb.active

            # 添加标题行
            column_names = list(testReport.keys())
            # 过滤掉None值
            column_names = [name for name in column_names if name is not None]
            for col_num, column_name in enumerate(column_names, 1):
                ws.cell(row=1, column=col_num, value=column_name)

            # 添加数据
            max_length = max(len(v) for v in testReport.values())  # 获取最长列的长度
            for row_num in range(2, max_length + 2):  # 从第二行开始添加数据
                for col_num, column_name in enumerate(column_names, 1):
                    if row_num - 2 < len(testReport[column_name]):
                        ws.cell(row=row_num, column=col_num, value=testReport[column_name][row_num - 2])

            # 保存Excel文件
            wb.save('power_strip_combined_data.xlsx')       
