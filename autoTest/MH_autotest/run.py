import os
import serial
import yaml
from MH_light_control import LightControl
import time
from datetime import datetime
from openpyxl import Workbook

# 获取当前路径并设置日志文件夹路径
current_path = os.getcwd()
now_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # 当前时间
result_path = os.path.join(current_path, f"MH控制器测试日志{now_time}")
if not os.path.exists(result_path):
    os.makedirs(result_path)
log_path = result_path
config_path = r"D:\meizhiguangdian\autoTest\MH_autotest\config\config.yaml"
with open(config_path, 'r', encoding='utf-8') as file:
    config = yaml.safe_load(file)
# 初始化LightControl实例
lightControl = LightControl(log_path)
pid = config['device']['pid']
ser_report = config['device']['ser']

def test_manual_light_adjust_mode(pid, start_dim):
    def perform_test(dimensions):
        for dim in dimensions:
            ser = serial.Serial(ser_report, 115200, timeout = 1)
            manual_dim = lightControl.manual_light_adjust_mode(ser, pid, 1, dim)
            if manual_dim is not None:
                ppfd_num = config['dimconfig'][f'dim{dim}']         
                difference = manual_dim - ppfd_num
                status = "通过" if -config['margins']['margin'] <= difference <= config['margins']['margin'] else "失败"
                log_msg = f"手动测试{status}: 设置{dim}%亮度，ppfd值为{manual_dim}, 差值为{difference}"
                lightControl.write_log(log_msg, 'test_report.txt')
            else:
                lightControl.write_log(f'manual_dim得值为：{manual_dim},请检查代码', 'test_report.txt')
            time.sleep(5)
    # 正向测试: 从start_dim到100，每步增加10%
    ascending_sequence = range(start_dim, 110, 10)
    perform_test(ascending_sequence)

    # 反向测试: 从100到start_dim，每步减少10%
    descending_sequence = range(100, start_dim-10, -10)
    perform_test(descending_sequence)


def test_auto_light_adjust_mode():
    st = config['autoParams']['st']
    et = config['autoParams']['et']
    darkenT = config['autoParams']['darkenT']
    offT = config['autoParams']['offT']
    ser = serial.Serial(ser_report, 115200, timeout=1)
    testReport = {}
    for sunTime in range(1, 21): # 循环遍历一个月中的每一天
        # 跨天
        if et >=1430:
            st = 0
            et = st + 2 * sunTime + 1
        lightControl.write_log(f'开始测试,开始时间为：{st},结束时间为：{et},日出日落时间为：{sunTime}', 'auto_light_mode.txt')
        data = lightControl.auto_light_adjust_mode(ser, pid, st, et, dim, darkenT, offT, sunTime)
        # 更新下一天的开始时间为今天的结束时间加1
        st = et + 1
        # 根据当前的sunTime计算当天的结束时间
        et = st + 2 * (sunTime + 1) + 1
        testReport[f'日出日落第{sunTime}分钟时间'] = data[1]
        testReport[f'日出日落第{sunTime}分钟ppfd值'] = data[0]

        print('-----------------------------------------------------------------')
        time.sleep(5)
    # 创建一个新的工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 添加标题行
    column_names = list(testReport.keys())
    for col_num, column_name in enumerate(column_names, 1):
        ws.cell(row=1, column=col_num, value=column_name)

    # 添加数据
    max_length = max(len(v) for v in testReport.values())  # 获取最长列的长度
    for row_num in range(2, max_length + 2):  # 从第二行开始添加数据
        for col_num, column_name in enumerate(column_names, 1):
            if row_num - 2 < len(testReport[column_name]):
                ws.cell(row=row_num, column=col_num, value=testReport[column_name][row_num - 2])

    # 保存Excel文件
    wb.save('combined_data.xlsx')

def test_ppfd_light_adjust_mod():
    st = config['ppfdParams']['st']
    et = config['ppfdParams']['et']
    dim_min = config['ppfdParams']['dim_min']
    dim_max = config['ppfdParams']['dim_max']
    ppfd = config['ppfdParams']['ppfd']
    darkenT = config['ppfdParams']['darkenT']
    offT = config['ppfdParams']['offT']
    sunTime = config['ppfdParams']['sunTime']
    ser = serial.Serial(ser_report, 115200, timeout=1)
    lightControl.ppfd_light_adjust_mode(ser, pid, st, et, dim_min, dim_max, ppfd, darkenT, offT, sunTime)


    

if __name__ == "__main__":
    dim = 10
    test_manual_light_adjust_mode(pid, dim)
    # test_ppfd_light_adjust_mod()